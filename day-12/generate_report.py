#!/usr/bin/env python3

import boto3
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import OrderedDict

EXCLUDE_SERVICES = {}

SERVICE_MAP = {
    'EC2 - Other': 'Amazon EC2',
    'Amazon Elastic Compute Cloud - Compute': 'Amazon EC2',
    'Amazon Elastic Block Store': 'Amazon EC2',
    'Amazon Elastic File System': 'Amazon EC2'
}

def get_cost_data(start, end):
    ce = boto3.client('ce')
    return ce.get_cost_and_usage(
        TimePeriod={'Start': start, 'End': end},
        Granularity='MONTHLY',
        Metrics=['UnblendedCost'],
        GroupBy=[{'Type': 'DIMENSION', 'Key': 'SERVICE'}]
    )

def write_excel_report(data, all_columns, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "AWS Cost Report"

    final_columns = ["Period"] + all_columns + ["Total costs($)"]
    ws.append(final_columns)

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill

    sorted_data = OrderedDict(sorted(
        data.items(),
        key=lambda item: datetime.strptime(item[0], '%Y-%m')
    ))

    for period, period_data in sorted_data.items():
        row = [period]
        total = 0.0
        for col in all_columns:
            #value = round(period_data.get(col, 0.0), 2)
            value = period_data.get(col, 0.0)
            row.append(value)
            total += value
        row.append(round(total, 2))
        ws.append(row)

    # Totals row
    total_row = ["TOTAL"]
    for col in all_columns:
        total_sum = sum(data[period].get(col, 0.0) for period in data)
        total_row.append(round(total_sum, 2))
    grand_total = sum(total_row[1:])
    total_row.append(round(grand_total, 2))
    ws.append(total_row)

    # Styling
    number_format = '#,##0.00'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = number_format
            if cell.row == 1:
                continue
            elif cell.row == ws.max_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    ws.freeze_panes = 'A2'
    wb.save(filename)
    print(f"\nReport saved to {filename}")

def main():
    end = datetime.today().replace(day=1)
    start = (end - timedelta(days=180)).replace(day=1)

    result = get_cost_data(start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d'))

    service_totals = {}
    monthly_data = {}

    for month_data in result['ResultsByTime']:
        month = month_data['TimePeriod']['Start'][:7]
        monthly_data[month] = {}
        for group in month_data['Groups']:
            raw_service = group['Keys'][0]
            service = SERVICE_MAP.get(raw_service, raw_service)
            if service in EXCLUDE_SERVICES:
                continue

            amount = float(group['Metrics']['UnblendedCost']['Amount'])
            if amount == 0:
                continue

            monthly_data[month][service] = monthly_data[month].get(service, 0.0) + amount
            service_totals[service] = service_totals.get(service, 0.0) + amount

    # Sort services by total cost (desc)
    sorted_services = sorted(service_totals.keys(), key=lambda s: service_totals[s], reverse=True)

    write_excel_report(monthly_data, sorted_services, 'aws_cost_report.xlsx')

if __name__ == "__main__":
    main()